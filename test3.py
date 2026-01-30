import boto3
import pandas as pd
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
import io
import os
import urllib.parse
import logging

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client('s3')

# --- Helper Functions from your original script ---
def ratio(x, y):
    try:
        return round(x / y * 100, 2)
    except ZeroDivisionError:
        return 0

def conf_list(df, conf_name):
    """Get distinct list from dataframe column"""
    raw_list = df[conf_name]
    return set(list(raw_list))

def generate_excel_report(bucket, key):
    logger.info(f"Downloading {key} from {bucket}")
    
    # 1. Read CSV from S3
    response = s3.get_object(Bucket=bucket, Key=key)
    csv_content = response['Body'].read()
    
    # Load into Pandas
    df = pd.read_csv(io.BytesIO(csv_content))
    
    # 2. Prepare Excel Output
    output = io.BytesIO()
    
    # Write initial data to Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=None, header=True)
    
    # Load workbook for formatting
    output.seek(0)
    book = openpyxl.load_workbook(output)
    
    # 3. Create 'Report' Sheet (Your Logic)
    if 'Report' not in book.sheetnames:
        book.create_sheet('Report')
    ws = book['Report']
    
    # Headers
    headers = ['ConfigRule', 'Total', 'Compliant', '%', 'Prod-Com', 'Prod-Non', 'P%', 'Stg-Com', 'Stg-Non', 'S%', 'Dev-Com', 'Dev-Non', 'D%']
    ws.append(headers)
    
    # Get distinct config rules
    distinct_rules = conf_list(df, "ConfigRule")
    
    # Calculate Stats
    for con in distinct_rules:
        # Filter main dataframe for this rule
        rule_df = df[df['ConfigRule'] == con]
        
        total = len(rule_df)
        compliant = len(rule_df[rule_df['ComplianceType'] == 'COMPLIANT'])
        tc_ratio = ratio(compliant, total)
        
        # Production Stats
        prod_df = rule_df[rule_df['AccountAlias'].str.contains("prod", na=False)]
        prod_com = len(prod_df[prod_df['ComplianceType'] == 'COMPLIANT'])
        prod_non = len(prod_df[prod_df['ComplianceType'] == 'NON_COMPLIANT'])
        pp_ratio = ratio(prod_com, (prod_com + prod_non))
        
        # Staging Stats
        stg_df = rule_df[rule_df['AccountAlias'].str.contains("stg", na=False)]
        stg_com = len(stg_df[stg_df['ComplianceType'] == 'COMPLIANT'])
        stg_non = len(stg_df[stg_df['ComplianceType'] == 'NON_COMPLIANT'])
        ss_ratio = ratio(stg_com, (stg_com + stg_non))
        
        # Dev Stats
        dev_df = rule_df[rule_df['AccountAlias'].str.contains("dev", na=False)]
        dev_com = len(dev_df[dev_df['ComplianceType'] == 'COMPLIANT'])
        dev_non = len(dev_df[dev_df['ComplianceType'] == 'NON_COMPLIANT'])
        dd_ratio = ratio(dev_com, (dev_com + dev_non))
        
        # Append Row
        ws.append([con, total, compliant, tc_ratio, prod_com, prod_non, pp_ratio, stg_com, stg_non, ss_ratio, dev_com, dev_non, dd_ratio])

    # 4. Apply Conditional Formatting
    # (Simplified range 'D2:D200' - ideally calculate max row dynamically, but this matches your script)
    color_red = 'C0504D'
    color_yellow = 'F79646'
    color_green = '9BBB59'
    
    rule = ColorScaleRule(start_type='percentile', start_value=10, start_color=color_red,
                          mid_type='percentile', mid_value=50, mid_color=color_yellow,
                          end_type='percentile', end_value=90, end_color=color_green)
    
    # Apply to percentage columns (D, G, J, M)
    for col in ['D', 'G', 'J', 'M']:
        ws.conditional_formatting.add(f'{col}2:{col}200', rule)

    # 5. Save back to S3
    output.seek(0)
    new_key = key.replace('.csv', '_Report.xlsx')
    s3.put_object(Bucket=bucket, Key=new_key, Body=output.getvalue())
    logger.info(f"Report uploaded to s3://{bucket}/{new_key}")

def lambda_handler(event, context):
    try:
        bucket = event['Records'][0]['s3']['bucket']['name']
        key = urllib.parse.unquote_plus(event['Records'][0]['s3']['object']['key'], encoding='utf-8')
        
        generate_excel_report(bucket, key)
        return "Success"
    except Exception as e:
        logger.error(f"Error: {e}")
        raise e